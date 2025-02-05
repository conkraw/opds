import pandas as pd
import streamlit as st
from openpyxl import load_workbook
from io import BytesIO
import matplotlib.pyplot as plt
import numpy as np

# Function to process each area based on the date column and starting row
def process_area(sheet, date_column, date_row, start_row, location):
    date = sheet[f'{date_column}{date_row}'].value
    am_names = [sheet[f'{date_column}{i}'].value for i in range(start_row, start_row + 10)]
    pm_names = [sheet[f'{date_column}{i}'].value for i in range(start_row + 10, start_row + 20)]
    names = [(name, 'AM') for name in am_names] + [(name, 'PM') for name in pm_names]

    processed_data = []
    for name, type_ in names:
        if name:
            preceptor, student = (name.split(' ~ ') if ' ~ ' in name else (name, None))
            student_placed = 'Yes' if student else 'No'
            student_type = None
            if student:
                if '(MD)' in student:
                    student_type = 'MD'
                elif '(PA)' in student:
                    student_type = 'PA'

            processed_data.append({
                'Date': date,
                'Type': type_,
                'Description': name,
                'Preceptor': preceptor.strip(),
                'Student': student.strip() if student else None,
                'Student Placed': student_placed,
                'Student Type': student_type,
                'Location': location
            })
    return processed_data

def correct_and_flag_missing_designations(df):
    """
    Corrects missing (MD) or (PA) designations in the dataset and flags rows where corrections were made.
    """
    df['Student Designation'] = df['Student'].str.extract(r'\((MD|PA)\)')
    df['Base Student Name'] = df['Student'].str.replace(r'\s*\(.*?\)', '', regex=True).str.strip()
    students_with_designations = df.dropna(subset=['Student Designation'])
    df['Correction Note'] = None

    for idx, row in df.iterrows():
        if row['Student Placed'] == 'Yes' and pd.isna(row['Student Designation']) and row['Base Student Name']:
            matches = students_with_designations[
                students_with_designations['Base Student Name'] == row['Base Student Name']
            ]
            if not matches.empty:
                correct_designation = matches['Student Designation'].iloc[0]
                corrected_student = f"{row['Base Student Name']} ({correct_designation})"
                df.at[idx, 'Student'] = corrected_student
                df.at[idx, 'Student Type'] = correct_designation  # Update Student Type
                df.at[idx, 'Correction Note'] = f"Corrected to '{corrected_student}'"

    df.drop(columns=['Student Designation', 'Base Student Name'], inplace=True)
    return df

st.title('OPD Data Processor')

uploaded_files = st.file_uploader("Choose Excel files", type="xlsx", accept_multiple_files=True)

if uploaded_files:
    all_data = []
    for uploaded_file in uploaded_files:
        wb = load_workbook(uploaded_file)
        for sheet_name in wb.sheetnames:
            sheet = wb[sheet_name]
            for date_column in ['B', 'C', 'D', 'E', 'F', 'G', 'H']:
                for date_row in [4, 28, 52, 76]:
                    start_row = date_row + 2
                    area_data = process_area(sheet, date_column, date_row, start_row, sheet_name)
                    all_data.extend(area_data)

    df = pd.DataFrame(all_data)

    # Exclude rows with "COM CLOSED" or "Closed" in the Description column
    df = df[~df['Description'].str.contains('COM CLOSED|Closed', case=False, na=False)]

    # Correct missing student designations and flag corrections
    df = correct_and_flag_missing_designations(df)

    # Add weekday column
    df['Date'] = pd.to_datetime(df['Date'], errors='coerce')
    df['Weekday'] = df['Date'].dt.day_name()

    # Calculate Days Worked by Preceptor
    filtered_df = df[df['Student Placed'] == 'Yes']

    # Filter for specific locations: HOPE_DRIVE, ETOWN, NYES
    filtered_locations = filtered_df[filtered_df['Location'].isin(['HOPE_DRIVE', 'ETOWN', 'NYES'])]
    
    # Count shifts assigned per student
    shifts_per_student = filtered_locations.groupby('Student').size().reset_index(name='Assigned Shifts')
    
    # Merge student type (MD/PA) for categorization
    shifts_per_student = pd.merge(shifts_per_student, filtered_locations[['Student', 'Student Type']].drop_duplicates(), on='Student', how='left')
    
    # Calculate averages
    average_shifts_total = shifts_per_student['Assigned Shifts'].mean()
    average_shifts_md = shifts_per_student[shifts_per_student['Student Type'] == 'MD']['Assigned Shifts'].mean()
    average_shifts_pa = shifts_per_student[shifts_per_student['Student Type'] == 'PA']['Assigned Shifts'].mean()
    
    # Display results in Streamlit
    st.write("Average Outpatient Shifts Assigned Per Individual Student (HOPE_DRIVE, ETOWN, NYES):")
    st.write(f"**Total Average Shifts**: {average_shifts_total:.2f}")
    st.write(f"**MD Average Shifts**: {average_shifts_md:.2f}")
    st.write(f"**PA Average Shifts**: {average_shifts_pa:.2f}")

    filtered_df['Half Day'] = filtered_df['Type'].apply(lambda x: 0.5 if x in ['AM', 'PM'] else 0)
    days_worked = (
        filtered_df.groupby(['Preceptor', 'Date'])['Half Day']
        .sum()
        .reset_index()
        .rename(columns={'Half Day': 'Total Day Fraction'})
    )

    preceptor_days_summary = (
        days_worked.groupby('Preceptor')['Total Day Fraction']
        .sum()
        .reset_index()
        .rename(columns={'Total Day Fraction': 'Total Days'})
    )

    # Calculate available and used shifts
    available_shifts = (
        df.groupby(['Preceptor', 'Date', 'Type'])
        .size()
        .reset_index(name='Available Shifts')
    )
    available_shifts = (
        available_shifts.groupby('Preceptor')['Available Shifts']
        .sum()
        .reset_index()
    )

    used_shifts = (
        filtered_df.groupby(['Preceptor', 'Date', 'Type'])
        .size()
        .reset_index(name='Used Shifts')
    )
    used_shifts = (
        used_shifts.groupby('Preceptor')['Used Shifts']
        .sum()
        .reset_index()
    )

    # Merge shifts data for preceptor summary
    shifts_summary = pd.merge(available_shifts, used_shifts, on='Preceptor', how='left')
    shifts_summary['Used Shifts'] = shifts_summary['Used Shifts'].fillna(0)

    # Calculate percentage of used shifts per provider
    shifts_summary['Percentage Used Shifts'] = (
        (shifts_summary['Used Shifts'] / shifts_summary['Available Shifts']) * 100
    ).fillna(0)

    # Calculate number of MD shifts
    md_shifts = filtered_df[filtered_df['Student Type'] == 'MD'].groupby('Preceptor').size().reset_index(name='MD Shifts')
    shifts_summary = pd.merge(shifts_summary, md_shifts, on='Preceptor', how='left')
    shifts_summary['MD Shifts'] = shifts_summary['MD Shifts'].fillna(0)

    # Calculate number of PA shifts
    pa_shifts = filtered_df[filtered_df['Student Type'] == 'PA'].groupby('Preceptor').size().reset_index(name='PA Shifts')
    shifts_summary = pd.merge(shifts_summary, pa_shifts, on='Preceptor', how='left')
    shifts_summary['PA Shifts'] = shifts_summary['PA Shifts'].fillna(0)

    # Calculate percentage of MD students
    shifts_summary['Percentage MD Students'] = (
        (shifts_summary['MD Shifts'] / shifts_summary['Used Shifts']) * 100
    ).fillna(0)

    # Plot total days worked
    fig, ax = plt.subplots(figsize=(10, 6))
    ax.bar(preceptor_days_summary['Preceptor'], preceptor_days_summary['Total Days'])
    ax.set_xlabel('Preceptor')
    ax.set_ylabel('Total Days Worked')
    ax.set_title('Total Days Worked by Preceptor')
    plt.xticks(rotation=45, ha='right', fontsize=10)
    st.pyplot(fig)

    # Plot available vs. used shifts
    fig, ax = plt.subplots(figsize=(12, 6))
    ax.bar(shifts_summary['Preceptor'], shifts_summary['Available Shifts'], label='Available Shifts', alpha=0.7)
    ax.bar(shifts_summary['Preceptor'], shifts_summary['Used Shifts'], label='Used Shifts', alpha=0.7)
    ax.set_xlabel('Preceptor')
    ax.set_ylabel('Shifts')
    ax.set_title('Available vs. Used Shifts by Preceptor')
    ax.legend()
    plt.xticks(rotation=45, ha='right', fontsize=10)
    st.pyplot(fig)

    # Plot percentage used shifts per provider
    fig, ax = plt.subplots(figsize=(12, 6))
    ax.bar(shifts_summary['Preceptor'], shifts_summary['Percentage Used Shifts'])
    ax.set_xlabel('Preceptor')
    ax.set_ylabel('Percentage of Used Shifts')
    ax.set_title('Percentage of Used Shifts Per Provider')
    plt.xticks(rotation=45, ha='right', fontsize=10)
    st.pyplot(fig)

    # Plot percentage MD students per provider
    fig, ax = plt.subplots(figsize=(12, 6))
    ax.bar(shifts_summary['Preceptor'], shifts_summary['Percentage MD Students'])
    ax.set_xlabel('Preceptor')
    ax.set_ylabel('Percentage of MD Students')
    ax.set_title('Percentage of MD Students Per Provider')
    plt.xticks(rotation=45, ha='right', fontsize=10)
    st.pyplot(fig)

    # Include all data in the downloadable Excel file
    output_file = BytesIO()
    with pd.ExcelWriter(output_file, engine='xlsxwriter') as writer:
        df.to_excel(writer, index=False, sheet_name='Combined Dataset')  # Full dataset
        preceptor_days_summary.to_excel(writer, index=False, sheet_name='Total Days Worked')  # Total days worked
        shifts_summary.to_excel(writer, index=False, sheet_name='Shifts Summary')  # Shifts summary with MD/PA shifts
    output_file.seek(0)

    st.download_button(
        label="Download Combined and Summary Data",
        data=output_file,
        file_name="combined_and_summary_data_with_pa_shifts.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

